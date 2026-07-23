"""
Lógica central de consolidación, reutilizada tanto por el script de terminal
como por la app web (Streamlit). No depende de la interfaz — solo recibe
rutas de archivo (o archivos subidos) y devuelve el resultado.
"""
import os
import shutil
import tempfile
import openpyxl

from xlsx_dropdown_fix import restore_dropdowns

MATRIZ_SHEET = "Matriz"
PLANTILLA_SHEET = "Solicitudes"
FIRST_DATA_ROW = 2
LAST_DATA_ROW = 101
KEY_COL = "O"


def leer_plantilla(file_or_path, nombre):
    wb = openpyxl.load_workbook(file_or_path, data_only=True)
    if PLANTILLA_SHEET not in wb.sheetnames:
        return [], f"'{nombre}' no tiene una hoja '{PLANTILLA_SHEET}' — se omitió."
    ws = wb[PLANTILLA_SHEET]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        n_interno, fecha_recep, distrito, descripcion, destino, fecha_contacto, cite, fecha_envio, fecha_resp, estado, obs = (
            (c.value if c else None) for c in row[:11]
        )
        if not distrito and not descripcion:
            continue
        filas.append({
            "n_interno": n_interno, "fecha_recepcion": fecha_recep, "distrito": distrito,
            "descripcion": descripcion, "destino": destino, "fecha_contacto": fecha_contacto,
            "cite": cite, "fecha_envio_formal": fecha_envio, "fecha_respuesta_real": fecha_resp,
            "estado": estado, "observaciones": obs,
        })
    return filas, None


def claves_existentes(ws):
    existentes = set()
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        v = ws[f"{KEY_COL}{r}"].value
        if v:
            existentes.add(v)
    return existentes


def primera_fila_libre(ws):
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        if not ws[f"C{r}"].value:
            return r
    return None


def intentar_recalcular(path):
    if not shutil.which("soffice") and not shutil.which("libreoffice"):
        return False
    recalc_candidates = [
        "/mnt/skills/public/xlsx/scripts/recalc.py",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "recalc.py"),
    ]
    recalc_script = next((p for p in recalc_candidates if os.path.exists(p)), None)
    if not recalc_script:
        return False
    os.system(f'python3 "{recalc_script}" "{path}" 60 >/dev/null 2>&1')
    restore_dropdowns(path, MATRIZ_SHEET)
    return True


def consolidar(matriz_path_or_file, plantillas, salida_path=None):
    """
    matriz_path_or_file: ruta o archivo en memoria de la Matriz maestra.
    plantillas: lista de tuplas (nombre_sin_extension, ruta_o_archivo_en_memoria).
    salida_path: si no se da, se sobrescribe matriz_path_or_file (debe ser ruta en ese caso).
    Devuelve: dict con nuevas, saltadas, sin_espacio, avisos (lista de str), resumen_filas (lista de dicts agregados).
    """
    wb = openpyxl.load_workbook(matriz_path_or_file)
    ws = wb[MATRIZ_SHEET]

    ya_importadas = claves_existentes(ws)
    nuevas = 0
    saltadas = 0
    sin_espacio = False
    avisos = []
    agregadas = []

    for nombre, fuente in plantillas:
        filas, aviso = leer_plantilla(fuente, nombre)
        if aviso:
            avisos.append(aviso)
            continue
        for f in filas:
            clave = f"{nombre}|{f['distrito']}|{f['n_interno']}"
            if clave in ya_importadas:
                saltadas += 1
                continue

            fila_libre = primera_fila_libre(ws)
            if fila_libre is None:
                sin_espacio = True
                avisos.append("La Matriz llegó a su límite de 100 filas. Hay solicitudes que no se pudieron importar.")
                break

            ws[f"B{fila_libre}"] = f["fecha_recepcion"]
            ws[f"C{fila_libre}"] = f["distrito"]
            ws[f"E{fila_libre}"] = f["descripcion"]
            ws[f"F{fila_libre}"] = f["destino"]
            ws[f"G{fila_libre}"] = f["fecha_contacto"]
            ws[f"H{fila_libre}"] = f["cite"]
            ws[f"I{fila_libre}"] = f["fecha_envio_formal"]
            ws[f"K{fila_libre}"] = f["fecha_respuesta_real"]
            ws[f"L{fila_libre}"] = f["estado"] or "Pendiente de Derivación"
            ws[f"N{fila_libre}"] = f["observaciones"]
            ws[f"{KEY_COL}{fila_libre}"] = clave

            ya_importadas.add(clave)
            nuevas += 1
            agregadas.append({"distrito": f["distrito"], "descripcion": f["descripcion"], "origen": nombre})
        if sin_espacio:
            break

    es_ruta = isinstance(matriz_path_or_file, str)
    salida_path = salida_path or (matriz_path_or_file if es_ruta else None)
    if salida_path is None:
        raise ValueError("Se necesita salida_path cuando la matriz de entrada no es una ruta de archivo")

    wb.save(salida_path)
    restore_dropdowns(salida_path, MATRIZ_SHEET)
    recalculo = intentar_recalcular(salida_path)

    return {
        "nuevas": nuevas,
        "saltadas": saltadas,
        "sin_espacio": sin_espacio,
        "avisos": avisos,
        "agregadas": agregadas,
        "recalculado": recalculo,
        "salida_path": salida_path,
    }


def consolidar_desde_bytes(matriz_bytes, plantillas_bytes):
    """
    matriz_bytes: bytes del archivo Matriz maestra subido.
    plantillas_bytes: lista de tuplas (nombre_archivo, bytes) de las plantillas subidas.
    Devuelve: (bytes_resultado, resumen_dict)
    """
    tmpdir = tempfile.mkdtemp(prefix="consolidacion_")
    try:
        matriz_path = os.path.join(tmpdir, "matriz.xlsx")
        with open(matriz_path, "wb") as f:
            f.write(matriz_bytes)

        plantillas = []
        for i, (nombre_archivo, contenido) in enumerate(plantillas_bytes):
            nombre_base = os.path.splitext(os.path.basename(nombre_archivo))[0]
            p_path = os.path.join(tmpdir, f"plantilla_{i}.xlsx")
            with open(p_path, "wb") as f:
                f.write(contenido)
            plantillas.append((nombre_base, p_path))

        salida_path = os.path.join(tmpdir, "matriz_consolidada.xlsx")
        resumen = consolidar(matriz_path, plantillas, salida_path=salida_path)

        with open(salida_path, "rb") as f:
            resultado_bytes = f.read()

        return resultado_bytes, resumen
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
